from openpyxl import load_workbook


usersFile = load_workbook("users.xlsx")


class User:
    def __init__(self, userID):
        userID = str(userID)
        self.userID = userID
        if userInDB(userID):
            self.userSheet = usersFile[self.userID]
        else:
            usersFile.create_sheet(self.userID)
            self.userSheet = usersFile[self.userID]
            usersFile.save("users.xlsx")
            print(f"Добавлен пользователь с ID {userID}")

    def addACameraIDToTheUser(self, cameraID):
        for i in range(1, self.userSheet.max_row + 1):
            if cameraID == self.userSheet.cell(row=i, column=1).value:
                return False
        else:
            self.userSheet.cell(row=self.userSheet.max_row + 1, column=1).value = cameraID
            usersFile.save("users.xlsx")
        return True

    def addCameraNameToTheUser(self, cameraName):
        self.userSheet.cell(row=self.userSheet.max_row, column=2).value = cameraName
        usersFile.save("users.xlsx")

    def getUserCameras(self):
        userCameras = []
        for i in range(1, self.userSheet.max_row + 1):
            userCameras.append(self.userSheet.cell(row=i, column=2).value)
        return userCameras

    def getUserCameraID(self, name):
        for i in range(1, self.userSheet.max_row + 1):
            if self.userSheet.cell(row=i, column=2).value == name:
                return self.userSheet.cell(row=i, column=1).value

    def getUserAddresses(self):
        inData = self.getUserCameras()
        outData = []
        for i in range(len(inData)):
            if inData[i]:
                outData.append([inData[i]])
        return outData


def userInDB(userID):
    userID = str(userID)
    sheets = usersFile.get_sheet_names()
    for user in sheets:
        if user == userID:
            return True
    return False


def addUserToBD(userID):
    userID = str(userID)
    usersFile.create_sheet(userID)
    usersFile.save("users.xlsx")
    print(f"Добавлен пользователь с ID {userID}")
