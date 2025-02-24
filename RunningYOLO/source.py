import cv2
import numpy as np
from openpyxl import load_workbook
import time

data_base = load_workbook("dataBase.xlsx")
camera_Pac = []
camera_count = 1


def nexStep():
    camera_Pac.cell(row=1, column=1).value += 1
    data_base.save("dataBase.xlsx")


def createData(annot_lines):
    if camera_Pac.cell(row=1, column=1).value == 0:
        m = camera_Pac.max_row - 1
        for parkingInd in range(m, len(annot_lines) + m):
            camera_Pac.cell(row=parkingInd + 2, column=1).value = annot_lines[parkingInd - m][0]
            camera_Pac.cell(row=parkingInd + 2, column=2).value = annot_lines[parkingInd - m][1]
            camera_Pac.cell(row=parkingInd + 2, column=3).value = annot_lines[parkingInd - m][2]
            camera_Pac.cell(row=parkingInd + 2, column=4).value = annot_lines[parkingInd - m][3]
            camera_Pac.cell(row=parkingInd + 2, column=5).value = 1
            camera_Pac.cell(row=parkingInd + 2, column=6).value = 0
        camera_Pac.cell(row=1, column=1).value += 1
        data_base.save("dataBase.xlsx")
        return True
    return False


def setCameraPac(newCamera):
    global camera_Pac
    camera_Pac = data_base[newCamera]


def delete_data():
    camera_Pac.delete_rows(1, camera_Pac.max_row)
    # for i in range(camera_Pac.max_row):
    #     camera_Pac.cell(row=i + 2, column=1).value = None
    #     camera_Pac.cell(row=i + 2, column=2).value = None
    #     camera_Pac.cell(row=i + 2, column=3).value = None
    #     camera_Pac.cell(row=i + 2, column=4).value = None
    #     camera_Pac.cell(row=i + 2, column=5).value = None
    camera_Pac.cell(row=1, column=1).value = 0
    data_base.save("dataBase.xlsx")


def now_all_space_free():
    tO = time.time()
    for i in range(2, camera_Pac.max_row + 1):
        if camera_Pac.cell(row=i, column=5).value and camera_Pac.cell(row=i, column=5).value != -1:
            camera_Pac.cell(row=i, column=6).value = 1
    data_base.save("dataBase.xlsx")
    tN = time.time()
    print("now_all_space_free", tN - tO)


def cheсk_free_space():
    free_space = []
    shlak_but_space = []
    not_free_space = []
    tO = time.time()
    for i in range(2, camera_Pac.max_row + 1):
        place = [
        camera_Pac.cell(row=i, column=1).value,
        camera_Pac.cell(row=i, column=2).value,
        camera_Pac.cell(row=i, column=3).value,
        camera_Pac.cell(row=i, column=4).value,
        camera_Pac.cell(row=i, column=5).value,
        camera_Pac.cell(row=i, column=6).value
    ]
        if place[5] == 1 and place[4] > 5:
            free_space.append(place)
        elif place[5] == 1 and place[4] <= 5:
            shlak_but_space.append(place)
        elif place[5] == 0:
            not_free_space.append(place)
    tN = time.time()
    print("cheсk_free_space", tN - tO, len(free_space))
    return free_space, shlak_but_space, not_free_space


def reduced_reliability():
    i = camera_Pac.max_row
    while i > 0:
        if camera_Pac.cell(row=i, column=6).value == 1:
            camera_Pac.cell(row=i, column=5).value -= 0.5
            if camera_Pac.cell(row=i, column=5).value < 0:
                camera_Pac.delete_rows(i)
        i -= 1


def same_box(box1, box2):
    n = 0
    for i in range(4):
        n += abs(box1[i] - box2[i])
    return n < 30


def delete_shit_in_data():
    i = 2
    max_row = camera_Pac.max_row
    while i <= max_row - 1:
        j = i + 1
        while j <= max_row:
            box1 = [
                camera_Pac.cell(row=i, column=1).value,
                camera_Pac.cell(row=i, column=2).value,
                camera_Pac.cell(row=i, column=3).value,
                camera_Pac.cell(row=i, column=4).value
                    ]
            box2 = [
                camera_Pac.cell(row=j, column=1).value,
                camera_Pac.cell(row=j, column=2).value,
                camera_Pac.cell(row=j, column=3).value,
                camera_Pac.cell(row=j, column=4).value
            ]
            if same_box(box1, box2):
                camera_Pac.delete_rows(j)
                max_row -= 1
                # camera_Pac.cell(row=i, column=1).value = None
                # camera_Pac.cell(row=i, column=2).value = None
                # camera_Pac.cell(row=i, column=3).value = None
                # camera_Pac.cell(row=i, column=4).value = None
                # camera_Pac.cell(row=i, column=5).value = None
                # camera_Pac.cell(row=i, column=6).value = None
                print("One more shit was deleted ", camera_Pac.cell(row=i, column=5).value, camera_Pac.cell(row=i, column=6).value)
                data_base.save("dataBase.xlsx")
            else:
                j += 1
        i += 1


def get_data():
    boxes = []
    for i in range(2, camera_Pac.max_row + 1):
        boxes.append([camera_Pac.cell(row=i, column=1).value,
                      camera_Pac.cell(row=i, column=2).value,
                      camera_Pac.cell(row=i, column=3).value,
                      camera_Pac.cell(row=i, column=4).value,
                      camera_Pac.cell(row=i, column=5).value,
                      camera_Pac.cell(row=i, column=6).value])

    return boxes


#Функции для подсчета Intersection over Union (IoU)
def calculate_iou(box):
    #Считаем IoU
    t = 0

    flag = 1
    for i in range(2, camera_Pac.max_row + 1):
        place = [
            camera_Pac.cell(row=i, column=1).value,
            camera_Pac.cell(row=i, column=2).value,
            camera_Pac.cell(row=i, column=3).value,
            camera_Pac.cell(row=i, column=4).value,
            camera_Pac.cell(row=i, column=5).value
        ]


        y1 = np.maximum(box[0], place[0])
        y2 = np.minimum(box[2]+box[0], place[2]+place[0])
        x1 = np.maximum(box[1], place[1])
        x2 = np.minimum(box[3]+box[1], place[3]+place[1])
        intersection = np.maximum(x2 - x1, 0) * np.maximum(y2 - y1, 0)
        union = box[2]*box[3] + place[2]*place[3] - intersection
        iou = intersection / union
        if iou > 0.6:
            #draw_two_box(image_to_process, [box, boxes[i]])
            #print(iou[i], box, boxes[i])
            midle = finde_midle(box, place)
            tO = time.time()
            camera_Pac.cell(row=i, column=1).value = midle[0]
            camera_Pac.cell(row=i, column=2).value = midle[1]
            camera_Pac.cell(row=i, column=3).value = midle[2]
            camera_Pac.cell(row=i, column=4).value = midle[3]
            camera_Pac.cell(row=i, column=5).value = midle[4]
            camera_Pac.cell(row=i, column=6).value = 0
            tN = time.time()
            t += tO - tN
            data_base.save("dataBase.xlsx")
            flag = 0
        if iou > 0.2:
            # midle = finde_midle(box, place)
            # midle = finde_midle(midle, place)
            # camera_Pac.cell(row=i, column=1).value = midle[0]
            # camera_Pac.cell(row=i, column=2).value = midle[1]
            # camera_Pac.cell(row=i, column=3).value = midle[2]
            # camera_Pac.cell(row=i, column=4).value = midle[3]
            # camera_Pac.cell(row=i, column=5).value = midle[4]
            camera_Pac.cell(row=i, column=6).value = 0
            data_base.save("dataBase.xlsx")
            flag = 0
    if flag:
        tO = time.time()
        row = camera_Pac.max_row + 1
        camera_Pac.cell(row=row, column=1).value = box[0]
        camera_Pac.cell(row=row, column=2).value = box[1]
        camera_Pac.cell(row=row, column=3).value = box[2]
        camera_Pac.cell(row=row, column=4).value = box[3]
        camera_Pac.cell(row=row, column=5).value = 1
        camera_Pac.cell(row=row, column=6).value = 0
        tN = time.time()
        t += tO - tN
    tO = time.time()
    data_base.save("dataBase.xlsx")
    tN = time.time()
    t += tO - tN
    #print("iou", t)
    return iou

#Функция для расчета персечения всех со всеми через IoU
def compute_overlaps(boxes1, boxes2, image_to_process):
    #Areas of anchors and GT boxes
    #print(boxes1, "\n", boxes2)
    # area1 = boxes1[:, 2] * boxes1[:, 3]
    # area2 = boxes2[:, 2] * boxes2[:, 3]
    # overlaps = np.zeros((boxes1.shape[0], boxes2.shape[0]))
    # for i in range(overlaps.shape[1]):
    #     box2 = boxes2[i]
    #     overlaps[:, i] = calculate_iou(box2, boxes1, area2[i], area1, image_to_process)

    for box in boxes2:
        calculate_iou(box)
    return


def finde_midle(box1, box2):
    new_box = [(box1[0] + box2[0]) / 2, (box1[1] + box2[1]) / 2, (box1[2] + box2[2]) / 2, (box1[3] + box2[3]) / 2, box2[4] + 1]
    return new_box


# Функция для отрисовки Bounding Box в кадре
def draw_bbox(box, image_to_process, parking_color=(0, 255, 0)):
    x, y, w, h = box[:4]
    start = (x, y)
    end = (x + w, y + h)
    color = parking_color
    width = 2
    final_image = cv2.rectangle(image_to_process, start, end, color, width)
    cv2.imshow('image', final_image)
    cv2.waitKey(0)

    # Подпись BB
    start = (x, y - 10)
    font_size = 0.4
    font = cv2.FONT_HERSHEY_SIMPLEX
    width = 1
    text = "parking_text"
    final_image = cv2.putText(final_image, text, start, font, font_size, color, width, cv2.LINE_AA)
    return final_image


def draw_data(image_to_process, boxes, parking_color=(0, 255, 0)):
    color = parking_color
    width = 2
    for box in boxes:
        # print(box)
        x, y, w, h = box[0], box[1], box[2], box[3]
        x, y, w, h = int(x), int(y), int(w), int(h)
        start = (x, y)
        end = (x + w, y + h)
        if box[5] and box[5] > 5:
            color = (0, 255, 0)
        elif box[5] and box[4] <= 5:
            color = (0, 165, 255)
        else:
            color = (255, 0, 0)
        image_to_process = cv2.rectangle(image_to_process, start, end, color, width)
        # cv2.imshow('image', image_to_process)
    return image_to_process


def draw_two_box(image_to_process, boxes):
    final_image = image_to_process
    for box in boxes:
        #print(box)
        x, y, w, h = box[0], box[1], box[2], box[3]
        start = (x, y)
        end = (x + w, y + h)
        color = (0, 255, 0)
        width = 2
        #final_image = cv2.rectangle(final_image, start, end, color, width)
        cv2.imshow('image', cv2.rectangle(final_image, start, end, color, width))
    #cv2.imshow('image', image_to_process)
    cv2.waitKey(0)
