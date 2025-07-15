import aiosqlite
import cv2
import logging
import numpy as np
from openpyxl import load_workbook
import time
import asyncpg

logger = logging.getLogger(__name__)

_db_ = None
_db_f = None
camera = str()
camera_Pac = []
camera_count = 1


class Place:
    """
    Класс для работы с данными о парковочных местах/зонах.
    Обеспечивает взаимодействие с базой данных (Excel через openpyxl) и основные операции с местами.
    """

    def __init__(self, line):
        """
        Инициализация объекта места.

        Args:
            line (int): Номер строки в базе данных, соответствующей этому месту.
        """
        self.line = line  # Номер строки в БД
        self.x = int()  # X-координата места
        self.y = int()  # Y-координата места
        self.h = int()  # Высота места
        self.w = int()  # Ширина места
        self.free = bool()  # Статус свободно/занято
        self.count = int()  # Счетчик использования
        self.confidence = int()  # Уверенность детекции
        self.sub = int()  # Дополнительный параметр

    def get(self):
        """
        Возвращает все параметры места в виде списка.

        Returns:
            list: [x, y, w, h, count, free, confidence, sub]
        """
        return [self.x, self.y, self.w, self.h, self.count, self.free, self.confidence, self.sub]

    def set(self, place, count=-1, confidence=-1, free=False, sub=-1):
        """
        Устанавливает параметры места.

        Args:
            place (list): Основные параметры [x, y, w, h]
            count (int): Счетчик использования (по умолчанию -1)
            confidence (int): Уверенность детекции (по умолчанию -1)
            free (bool): Статус свободно (по умолчанию False)
            sub (int): Дополнительный параметр (по умолчанию -1)
        """
        self.x, self.y, self.w, self.h, self.count, self.confidence, self.free, self.sub = place[0], place[1], place[2], \
                                                                                           place[
                                                                                               3], count, confidence, free, sub

    async def load(self):
        """Загружает данные места из базы данных по номеру строки."""
        if self.line == 1 and _db_.cell(row=1, column=1).value == None:
            self.x = 0
            self.y = 0
            self.h = 0
            self.w = 0
            self.free = True
            self.count = 0
            self.confidence = 0
            self.sub = 0
        else:
            self.x = _db_.cell(row=self.line, column=1).value
            self.y = _db_.cell(row=self.line, column=2).value
            self.h = _db_.cell(row=self.line, column=3).value
            self.w = _db_.cell(row=self.line, column=4).value
            self.free = _db_.cell(row=self.line, column=5).value
            self.count = _db_.cell(row=self.line, column=6).value
            self.confidence = _db_.cell(row=self.line, column=7).value
            self.sub = _db_.cell(row=self.line, column=8).value

    async def push(self):
        """Сохраняет текущие параметры места в базу данных."""
        _db_.cell(row=self.line, column=1).value = self.x
        _db_.cell(row=self.line, column=2).value = self.y
        _db_.cell(row=self.line, column=3).value = self.h
        _db_.cell(row=self.line, column=4).value = self.w
        _db_.cell(row=self.line, column=5).value = self.free
        _db_.cell(row=self.line, column=6).value = self.count
        _db_.cell(row=self.line, column=7).value = self.confidence
        _db_.cell(row=self.line, column=8).value = self.sub

    async def add(self):
        """
        Добавляет новое место в конец базы данных.
        Автоматически определяет номер строки.
        """
        self.line = _db_.max_row + 1
        if _db_.cell(row=1, column=1).value == None:
            self.line = 1
        _db_.cell(row=self.line, column=1).value = self.x
        _db_.cell(row=self.line, column=2).value = self.y
        _db_.cell(row=self.line, column=3).value = self.h
        _db_.cell(row=self.line, column=4).value = self.w
        _db_.cell(row=self.line, column=5).value = self.free
        _db_.cell(row=self.line, column=6).value = self.count
        _db_.cell(row=self.line, column=7).value = self.confidence
        _db_.cell(row=self.line, column=8).value = self.sub

    def finde_midle(self, box):
        """
        Вычисляет среднее значение координат между текущим местом и переданным bounding box.

        Args:
            box (list): Bounding box [x, y, w, h]
        """
        self.x = (self.x + box[0]) / 2
        self.y = (self.y + box[1]) / 2
        self.w = (self.w + box[2]) / 2
        self.h = (self.h + box[3]) / 2
        self.count += 1  # Увеличиваем счетчик использований

    async def delete(self):
        """Удаляет запись о месте из базы данных."""
        _db_.delete_rows(self.line)


async def save():
    """Сохраняет текущее состояние базы данных (файл Excel)"""
    # print("saving...")
    _db_f.save(f"bd/{camera}.xlsx")
    # print("saving complete")


async def init_db(newCamera):
    """Загрухает файл текущей камеры"""
    global _db_
    global _db_f
    global camera
    camera = newCamera
    _db_f = load_workbook(f"bd/{camera}.xlsx")
    if _db_f:
        _db_ = _db_f["main"]
        print("successful connect")


async def now_all_space_free():
    """Помечает все места в базе данных как свободные"""
    for i in range(1, _db_.max_row + 1):
        _db_.cell(row=i, column=5).value = True


async def cheсk_free_space():
    """
    Анализирует и классифицирует все места в базе данных по трем категориям:
    1. Свободные места с высокой достоверностью (confidence > 5)
    2. Потенциально свободные места с низкой достоверностью
    3. Занятые места

    Returns:
        tuple: (free_space, shlak_but_space, not_free_space)
    """
    free_space = []
    shlak_but_space = []
    not_free_space = []
    for i in range(1, _db_.max_row + 1):
        place = Place(i)
        await place.load()
        if place.free and place.confidence > 5:
            free_space.append(place)
        elif place.free:
            shlak_but_space.append(place)
        elif not place.free:
            not_free_space.append(place)
    return free_space, shlak_but_space, not_free_space


async def reduced_reliability():
    """
    Уменьшает показатель достоверности для свободных мест.
    Если достоверность падает ниже 0 - место удаляется из базы.
    """
    for i in range(1, _db_.max_row + 1):
        place = Place(i)
        await place.load()
        if place.free:
            place.confidence -= 0.5
            if place.confidence < 0:
                await place.delete()
            else:
                await place.push()


async def calculate_iou(box, push=True):
    """
    Вычисляет Intersection over Union (IoU) между переданным bounding box и всеми местами в БД.
    Обновляет статус мест на основе результатов вычислений.

    Args:
        box (list): Bounding box в формате [x, y, width, height]
        push (bool): Флаг сохранения изменений в БД (по умолчанию True)

    Процесс работы:
    1. Для каждого места вычисляется IoU с переданным box
    2. Если IoU > 0.6 - место считается соответствующим box'у:
       - Координаты места усредняются с box
       - Увеличивается confidence
    3. Если IoU > 0.1 - место помечается как занятое и все
    4. Если ни одно место не соответствует box'у - создается новое место
    """
    t = 0
    totalIOU = 0
    maxIOU = 0
    flag = 1  # Флаг для определения необходимости создания нового места

    for i in range(1, _db_.max_row + 1):
        place = Place(i)
        await place.load()

        # Вычисление координат пересечения
        y1 = np.maximum(box[0], place.x)
        y2 = np.minimum(box[2] + box[0], place.w + place.x)
        x1 = np.maximum(box[1], place.y)
        x2 = np.minimum(box[3] + box[1], place.h + place.y)

        # Вычисление площади пересечения и объединения
        intersection = np.maximum(x2 - x1, 0) * np.maximum(y2 - y1, 0)
        union = box[2] * box[3] + place.h * place.w - intersection
        iou = intersection / union

        totalIOU += iou

        # Логика обновления места на основе IoU
        if iou > 0.6:  # Сильное пересечение
            place.finde_midle(box)  # Усредняем координаты
            tO = time.time()
            place.sub = iou
            place.confidence += 1  # Увеличиваем достоверность
            tN = time.time()
            t += tO - tN
            flag = 0  # Новое место не нужно

        if iou > 0.1:  # Среднее пересечение
            place.free = False  # Помечаем как занятое
            flag = 0
        else:
            if place.sub + iou > 0.1:
                place.free = False
            place.sub += iou

        if push:
            await place.push()  # Сохраняем изменения

    # Если ни с одним местом нет достаточного пересечения - создаем новое
    if flag:
        place = Place(1)
        place.set(box, 0, 5, False, 0)
        if push:
            await place.add()

    return True


async def compute_overlaps(boxes, newCamera):
    """
    Основная функция обработки bounding boxes:
    1. Открываем файл БД
    2. Сбрасывает все IoU значения
    3. Помечает все места как свободные
    4. Обрабатывает каждый bounding box
    5. Уменьшает достоверность свободных мест
    6. Классифицирует места
    7. Сохраняет изменения
    """
    await init_db(newCamera)
    await setIOU()
    await now_all_space_free()
    for box in boxes:
        place = Place(-1)
        place.set(box)
        await calculate_iou(box)
    await reduced_reliability()
    x = await cheсk_free_space()
    await save()
    return x


async def setIOU():
    """Сбрасывает все значения IoU в базе данных"""
    for i in range(1, _db_.max_row + 1):
        _db_.cell(row=i, column=8).value = 0


async def draw_data(image_to_process, free_space, shlak, not_free_space, parking_color=(0, 255, 0)):
    """
    Визуализирует места на изображении с цветовой кодировкой:
    - Зеленый: свободные с высокой достоверностью
    - Оранжевый: свободные с низкой достоверностью
    - Синие: занятые места
    """
    color = parking_color
    width = 2
    for place in free_space:
        if place.x and place.y and place.w and place.h:
            x, y, w, h = int(place.x), int(place.y), int(place.w), int(place.h)
            start = (x, y)
            end = (x + w, y + h)
            color = (0, 255, 0)
            image_to_process = cv2.rectangle(image_to_process, start, end, color, width)

    for place in shlak:
        if place.x and place.y and place.w and place.h:
            x, y, w, h = int(place.x), int(place.y), int(place.w), int(place.h)
            start = (x, y)
            end = (x + w, y + h)
            color = (0, 165, 255)
            image_to_process = cv2.rectangle(image_to_process, start, end, color, width)

    for place in not_free_space:
        if place.x and place.y and place.w and place.h:
            x, y, w, h = int(place.x), int(place.y), int(place.w), int(place.h)
            start = (x, y)
            end = (x + w, y + h)
            color = (255, 0, 0)
            image_to_process = cv2.rectangle(image_to_process, start, end, color, width)

    return image_to_process


async def delete_data():
    """Удаляет все данные из таблиц, но сохраняет структуру"""
    async with aiosqlite.connect('core/databases/users.db') as db:
        try:
            # Очищаем таблицу users
            await db.execute("DELETE FROM users")

            # Если есть другие таблицы, их тоже можно очистить
            # await db.execute("DELETE FROM other_table")

            await db.commit()
            logger.info('Все данные пользователей удалены')

        except Exception as e:
            await db.rollback()
            logger.error(f'Ошибка при удалении данных: {e}')

    """Удаляет все данные из subscriptions.db, сохраняя структуру таблиц"""
    async with aiosqlite.connect('core/databases/subscriptions.db') as db:
        try:
            # Получаем список всех таблиц в базе
            cursor = await db.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in await cursor.fetchall()]

            # Очищаем каждую таблицу
            for table in tables:
                await db.execute(f"DELETE FROM {table}")
                logger.info(f'Данные из таблицы {table} удалены')

            await db.commit()
            logger.info('Все данные в subscriptions.db успешно очищены')

        except Exception as e:
            await db.rollback()
            logger.error(f'Ошибка при очистке subscriptions.db: {e}')

    """Удаляет все данные из referrals.db, сохраняя структуру таблиц"""
    async with aiosqlite.connect('core/databases/referrals.db') as db:
        try:
            # Получаем список всех таблиц в базе
            cursor = await db.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in await cursor.fetchall()]

            # Очищаем каждую таблицу
            for table in tables:
                await db.execute(f"DELETE FROM {table}")
                logger.info(f'Данные из таблицы {table} удалены')

            await db.commit()
            logger.info('Все данные в referrals.db успешно очищены')

        except Exception as e:
            await db.rollback()
            logger.error(f'Ошибка при очистке referrals.db: {e}')



#
# import cv2
# import numpy as np
# from openpyxl import load_workbook
# import time
# import asyncpg
#
# _db_pool = None
# data_base = load_workbook("dataBase.xlsx")
# camera_Pac = []
# camera_count = 1
#
#
# class Place:
#     def __init__(self, line):
#         self.line = line
#         self.x = int()
#         self.y = int()
#         self.h = int()
#         self.w = int()
#         self.free = bool()
#         self.count = int()
#         self.confidence = int()
#         self.sub = int()
#
#     def get(self):
#         return [self.x, self.y, self.w, self.h, self.count, self.free, self.confidence, self.sub]
#
#     def set(self, place, count=-1, confidence=-1, free=False, sub=-1):
#         self.x, self.y, self.w, self.h, self.count, self.confidence, self.free, self.sub = place[0], place[1], place[2], \
#                                                                                            place[
#                                                                                                3], count, confidence, free, sub
#
#     async def load(self, camera):
#         tO = time.time()
#         async with _db_pool.acquire() as conn:
#             row = await conn.fetchrow(
#                 f"""SELECT "X", "Y", "H", "W", "COUNT", "CONFIDENCE", "FREE", "SUB" FROM parking."{camera}" WHERE ctid = '{self.line})'::tid"""
#             )
#             if row:
#                 self.x = row["X"]
#                 self.y = row["Y"]
#                 self.h = row["H"]
#                 self.w = row["W"]
#                 self.free = row["FREE"]
#                 self.count = row["COUNT"]
#                 self.confidence = row["CONFIDENCE"]
#                 self.sub = row["SUB"]
#                 tN = time.time()
#                 #print(camera, "load place", tN - tO, row)
#
#     async def push(self, camera):
#         async with _db_pool.acquire() as conn:
#             query = f"""
#             UPDATE parking."{camera}" SET
#                 "X" = {self.x},
#                 "Y" = {self.y},
#                 "H" = {self.h},
#                 "W" = {self.w},
#                 "COUNT" = {self.count},
#                 "CONFIDENCE" = {self.confidence},
#                 "FREE" = {self.free},
#                 "SUB" = {self.sub}
#             WHERE ctid = '{self.line}'::tid;
#             """
#             await conn.execute(query)
#             # await conn.execute(
#             #     f'''
#             #                 DO $$
#             #             BEGIN
#             #                 IF EXISTS (
#             #                     SELECT 1 FROM parking."{camera}"
#             #                     WHERE ctid = '(0,{self.line})'::tid
#             #                 ) THEN
#             #                     UPDATE parking."{camera}" SET
#             #                         "X" = {self.x},
#             #                         "Y" = {self.y},
#             #                         "H" = {self.h},
#             #                         "W" = {self.w},
#             #                         "COUNT" = {self.count},
#             #                         "CONFIDENCE" = {self.confidence},
#             #                         "FREE" = {self.free},
#             #                         "SUB" = {self.sub}
#             #                     WHERE ctid = '(0,{self.line})'::tid;
#             #                 ELSE
#             #                     INSERT INTO parking."{camera}"
#             #                         ("X", "Y", "H", "W", "COUNT", "CONFIDENCE", "FREE", "SUB")
#             #                     VALUES (
#             #                         {self.x}, {self.y}, {self.h}, {self.w},
#             #                         {self.count}, {self.confidence}, {self.free}, {self.sub}
#             #                     );
#             #                 END IF;
#             #             END $$;
#             #     ''',
#             #     self.x, self.y, self.h, self.w,
#             #     self.count, self.confidence, self.free, self.sub
#             # )
#
#     async def add(self, camera):
#         async with _db_pool.acquire() as conn:
#             query = f"""
#             INSERT INTO parking."{camera}"
#                 ("X", "Y", "H", "W", "COUNT", "CONFIDENCE", "FREE", "SUB")
#             VALUES ({self.x}, {self.y}, {self.h}, {self.w}, {self.count}, {self.confidence}, {self.free}, {self.sub})
#             """
#             await conn.execute(query)
#
#     def finde_midle(self, box):
#         self.x = (self.x + box[0]) / 2
#         self.y = (self.y + box[1]) / 2
#         self.w = (self.w + box[2]) / 2
#         self.h = (self.h + box[3]) / 2
#         self.count += 1
#
#     async def delete(self, camera):
#         async with _db_pool.acquire() as conn:
#             await conn.execute(
#                 f"""DELETE FROM parking."{camera}" WHERE ctid = '{self.line}'::tid"""
#             )
#
#
# async def shutdown():
#     await _db_pool.close()  # Запрещаем новые подключения
#
#
# async def init_db(camera):
#     global _db_pool
#     if camera == '':
#         _db_pool = await asyncpg.create_pool(
#             f'postgresql://parking_admin:ParkinG!23@185.250.44.14:5432/parking_db'
#         )
#     # 'postgresql://parking_admin:ParkinG!23@185.250.44.14:5432/parking_db'
#     else:
#         _db_pool = await asyncpg.create_pool(
#             f'postgresql://{camera}_admin:ParkinG!23@185.250.44.14:5432/parking_db'
#         )
#     try:
#         async with _db_pool.acquire() as conn:
#             await conn.fetch("SELECT 1")
#         print("successful connect")
#     except:
#         print("failed connect")
#
#
# async def now_all_space_free(camera):
#     async with _db_pool.acquire() as conn:
#         await conn.execute(
#             f'UPDATE parking."{camera}" SET "FREE" = True'
#         )
#
#
# async def cheсk_free_space(camera):
#     free_space = []
#     shlak_but_space = []
#     not_free_space = []
#     async with _db_pool.acquire() as conn:
#         rows = await conn.fetch(f"""
#                 SELECT ctid
#                 FROM parking."{camera}"
#                 ORDER BY ctid;
#             """)
#         print(rows)
#     for ctid in rows:
#         ctid = ctid['ctid']
#         place = Place(ctid)
#         await place.load(camera)
#         if place.free and place.confidence > 5:
#             free_space.append(place)
#         elif place.free:
#             shlak_but_space.append(place)
#         elif not place.free:
#             not_free_space.append(place)
#     return free_space, shlak_but_space, not_free_space
#
#
# async def reduced_reliability(camera):
#     async with _db_pool.acquire() as conn:
#         rows = await conn.fetch(f"""
#                 SELECT ctid
#                 FROM parking."{camera}"
#                 ORDER BY ctid;
#             """)
#         print(rows)
#     for ctid in rows:
#         ctid = ctid['ctid']
#         place = Place(ctid)
#         await place.load(camera)
#         if place.free:
#             place.confidence -= 0.5
#             if place.confidence < 0:
#                 await place.delete(camera)
#     async with _db_pool.acquire() as conn:
#         await conn.fetchval(f'VACUUM FULL parking."{camera}"')
#
#
# def same_box(box1, box2):
#     n = 0
#     for i in range(4):
#         n += abs(box1[i] - box2[i])
#     return n < 30
#
#
# def delete_shit_in_data():
#     i = 2
#     max_row = camera_Pac.max_row
#     while i <= max_row - 1:
#         j = i + 1
#         while j <= max_row:
#             box1 = [
#                 camera_Pac.cell(row=i, column=1).value,
#                 camera_Pac.cell(row=i, column=2).value,
#                 camera_Pac.cell(row=i, column=3).value,
#                 camera_Pac.cell(row=i, column=4).value
#             ]
#             box2 = [
#                 camera_Pac.cell(row=j, column=1).value,
#                 camera_Pac.cell(row=j, column=2).value,
#                 camera_Pac.cell(row=j, column=3).value,
#                 camera_Pac.cell(row=j, column=4).value
#             ]
#             if same_box(box1, box2):
#                 camera_Pac.delete_rows(j)
#                 max_row -= 1
#                 print("One more shit was deleted ", camera_Pac.cell(row=i, column=5).value,
#                       camera_Pac.cell(row=i, column=6).value)
#                 data_base.save("dataBase.xlsx")
#             else:
#                 j += 1
#         i += 1
#
#
# async def calculate_iou(box, camera, push=True):
#     t = 0
#     totalIOU = 0
#     maxIOU = 0
#     flag = 1
#
#     async with _db_pool.acquire() as conn:
#         rows = await conn.fetch(f"""
#         SELECT ctid
#         FROM parking."{camera}"
#         ORDER BY ctid;
#     """)
#         print(rows)
#     print(camera, box, end='\n')
#     for ctid in rows:
#         ctid = ctid['ctid']
#         place = Place(ctid)
#         await place.load(camera)
#         print(ctid, place.get(), end=' ')
#         y1 = np.maximum(box[0], place.x)
#         y2 = np.minimum(box[2] + box[0], place.w + place.x)
#         x1 = np.maximum(box[1], place.y)
#         x2 = np.minimum(box[3] + box[1], place.h + place.y)
#         intersection = np.maximum(x2 - x1, 0) * np.maximum(y2 - y1, 0)
#         union = box[2] * box[3] + place.h * place.w - intersection
#         iou = intersection / union
#         totalIOU += iou
#         print(iou, end=' ')
#         if iou > 0.6:
#             place.finde_midle(box)
#             tO = time.time()
#             place.sub = iou
#             place.confidence += 1
#             tN = time.time()
#             t += tO - tN
#             flag = 0
#             print("correct", end=" ")
#         if iou > 0.15:
#             place.free = False
#             flag = 0
#             print("mark", end=" ")
#         else:
#             if place.sub + iou > 0.15:
#                 place.free = False
#             place.sub += iou
#             print("free", end=" ")
#         print("\n")
#         if push:
#             await place.push(camera)
#     async with _db_pool.acquire() as conn:
#         await conn.fetchval(f'VACUUM FULL parking."{camera}"')
#
#     if flag:
#         place = Place(0)
#         place.set(box, 0, 5, False, 0)
#         if push:
#             #print("add")
#             await place.add(camera)
#     #print('\n')
#     return True
#
#
# async def getOne():
#     tables = await get_all_tables()
#     for table in tables:
#         if table != "CAMERAS":
#             async with _db_pool.acquire() as conn:
#                 rows = await conn.fetch(f'SELECT ctid, * FROM parking."{table}"')
#                 for row in rows:
#                     print(f"{table}--- CTID: {row['ctid']}, Data: {dict(row)}")
#
#
#
# async def compute_overlaps(boxes, camera):
#     await init_db(camera)
#     #print(await get_all_tables())
#     await setIOU(camera)
#     await now_all_space_free(camera)
#     for box in boxes:
#         place = Place(-1)
#         place.set(box)
#         await calculate_iou(box, camera)
#     await reduced_reliability(camera)
#     await getOne()
#     x = await cheсk_free_space(camera)
#     await shutdown()
#     return x
#
#
# async def setIOU(camera):
#     async with _db_pool.acquire() as conn:
#         await conn.execute(
#             f'UPDATE parking."{camera}" SET "SUB" = 0'
#         )
#
#
# async def draw_data(image_to_process, boxes, parking_color=(0, 255, 0)):
#     color = parking_color
#     width = 2
#     for i in boxes:
#         for place in i:
#             x, y, w, h = int(place.x), int(place.y), int(place.w), int(place.h)
#             start = (x, y)
#             end = (x + w, y + h)
#             if place.free and place.confidence > 5:
#                 color = (0, 255, 0)
#             elif place.free and place.confidence <= 5:
#                 color = (0, 165, 255)
#             else:
#                 color = (255, 0, 0)
#             image_to_process = cv2.rectangle(image_to_process, start, end, color, width)
#     return image_to_process
#
#
# async def get_all_tables():
#     async with _db_pool.acquire() as conn:
#         tables = await conn.fetch(
#             "SELECT table_name FROM information_schema.tables "
#             "WHERE table_schema = 'parking' AND table_type = 'BASE TABLE'"
#         )
#         return [table['table_name'] for table in tables]
#
#
# async def delete_data():
#     await init_db('')
#     tables = await get_all_tables()
#     for table in tables:
#         if table != "CAMERAS":
#             async with _db_pool.acquire() as conn:
#                 await conn.execute(f'DELETE FROM parking."{table}"')
#                 await conn.fetchval(f'VACUUM FULL parking."{table}"')
#




"""
import cv2
import numpy as np
from openpyxl import load_workbook
import time
import asyncpg

_db_pool = None
data_base = load_workbook("dataBase.xlsx")
camera_Pac = []
camera_count = 1


class Place:
    def __init__(self, line):
        self.line = line
        self.x = int()
        self.y = int()
        self.h = int()
        self.w = int()
        self.free = bool()
        self.count = int()
        self.confidence = int()
        self.sub = int()

    def get(self):
        return [self.x, self.y, self.h, self.w, self.count, self.free, self.confidence, self.sub]

    def set(self, place, count=-1, confidence=-1, free=False, sub=-1):
        self.x, self.y, self.h, self.w, self.count, self.confidence, self.free, self.sub = place[0], place[1], place[2], place[3], count, confidence, free, sub

    def load(self, camera):
        tO = time.time()
        async with _db_pool.acquire() as conn:
            row = await conn.fetchrow(
                f'SELECT x, y, h, w, count, confidence, free, sub FROM {camera} WHERE id = $1',
                self.line  # ID нужной строки
            )
            await conn.close()
            self.x = row['x']
            self.y = row['y']
            self.h = row['h']
            self.w = row['w']
            self.free = row['free']
            self.count = row['count']
            self.confidence = row['confidence']
            self.sub = row['sub']
            tN = time.time()
            print(camera, "load place", tN - tO)

    async def push(self, camera):
        async with _db_pool.acquire() as conn:
            await conn.execute(
                f'''
                INSERT INTO {camera} (id, x, y, h, w, count, confidence, free, sub)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                ON CONFLICT (id) DO UPDATE
                SET
                    x = EXCLUDED.x,
                    y = EXCLUDED.y,
                    h = EXCLUDED.h,
                    w = EXCLUDED.w,
                    count = EXCLUDED.count,
                    confidence = EXCLUDED.confidence,
                    free = EXCLUDED.free,
                    sub = EXCLUDED.sub
                ''',
                self.line, self.x, self.y, self.h, self.w,
                self.count, self.confidence, self.free, self.sub
            )

    def finde_midle(self, box):
        self.x = (self.x + box[0]) / 2
        self.y = (self.y + box[1]) / 2
        self.h = (self.h + box[2]) / 2
        self.w = (self.w + box[3]) / 2
        self.count += 1

    def delete(self, camera):
        async with _db_pool.acquire() as conn:
            row = await conn.fetchrow(
                f"DELETE FROM {camera} WHERE id = $1",
                self.line  # ID нужной строки
            )


async def init_db():
    global _db_pool
    #'postgresql://parking_admin:ParkinG!23@localhost/parking_db'
    _db_pool = await asyncpg.create_pool(
        'postgresql://parking_admin:ParkinG!23@185.250.44.14:5432/parking_db'
    )


def nexStep():
    camera_Pac.cell(row=1, column=1).value += 1
    data_base.save("dataBase.xlsx")


def createData(annot_lines):
    if camera_Pac.cell(row=1, column=1).value == 0:
        m = camera_Pac.max_row - 1
        for parkingInd in range(m, len(annot_lines) + m):
            camera_Pac.cell(row=parkingInd + 2, column=1).value = annot_lines[parkingInd - m][0]
            camera_Pac.cell(row=parkingInd + 2, column=2).value = annot_lines[parkingInd - m][1]
            camera_Pac.cell(row=parkingInd + 2, column=3).value = annot_lines[parkingInd - m][2]
            camera_Pac.cell(row=parkingInd + 2, column=4).value = annot_lines[parkingInd - m][3]
            camera_Pac.cell(row=parkingInd + 2, column=5).value = 1
            camera_Pac.cell(row=parkingInd + 2, column=6).value = 0
        camera_Pac.cell(row=1, column=1).value += 1
        data_base.save("dataBase.xlsx")
        return True
    return False


def setCameraPac(newCamera):
    global camera_Pac
    camera_Pac = data_base[newCamera]


def delete_data():
    camera_Pac.delete_rows(1, camera_Pac.max_row)
    # for i in range(camera_Pac.max_row):
    #     camera_Pac.cell(row=i + 2, column=1).value = None
    #     camera_Pac.cell(row=i + 2, column=2).value = None
    #     camera_Pac.cell(row=i + 2, column=3).value = None
    #     camera_Pac.cell(row=i + 2, column=4).value = None
    #     camera_Pac.cell(row=i + 2, column=5).value = None
    camera_Pac.cell(row=1, column=1).value = 0
    data_base.save("dataBase.xlsx")



def now_all_space_free(camera):
    async with _db_pool.acquire() as conn:
        await conn.execute(
            f"UPDATE {camera} SET free = True"
        )


def cheсk_free_space(camera):
    free_space = []
    shlak_but_space = []
    not_free_space = []
    async with _db_pool.acquire() as conn:
        lenth = await conn.fetchval(f"SELECT COUNT(*) FROM {camera}")
    for i in range(lenth):
        place = Place(i)
        place.load()
        if place.free and place.confidence > 5:
            free_space.append(place)
        elif place.free:
            shlak_but_space.append(place)
        elif not place.free:
            not_free_space.append(place)

    return free_space, shlak_but_space, not_free_space


def reduced_reliability(camera):
    async with _db_pool.acquire() as conn:
        i = await conn.fetchval(f"SELECT COUNT(*) FROM {camera}")
    while i > 0:
        place = Place(i)
        place.load()
        if place.free:
            place.confidence -= 0.5
            if place.confidence < 0:
                place.delete()
        i -= 1


def same_box(box1, box2):
    n = 0
    for i in range(4):
        n += abs(box1[i] - box2[i])
    return n < 30


def delete_shit_in_data():
    i = 2
    max_row = camera_Pac.max_row
    while i <= max_row - 1:
        j = i + 1
        while j <= max_row:
            box1 = [
                camera_Pac.cell(row=i, column=1).value,
                camera_Pac.cell(row=i, column=2).value,
                camera_Pac.cell(row=i, column=3).value,
                camera_Pac.cell(row=i, column=4).value
            ]
            box2 = [
                camera_Pac.cell(row=j, column=1).value,
                camera_Pac.cell(row=j, column=2).value,
                camera_Pac.cell(row=j, column=3).value,
                camera_Pac.cell(row=j, column=4).value
            ]
            if same_box(box1, box2):
                camera_Pac.delete_rows(j)
                max_row -= 1
                # camera_Pac.cell(row=i, column=1).value = None
                # camera_Pac.cell(row=i, column=2).value = None
                # camera_Pac.cell(row=i, column=3).value = None
                # camera_Pac.cell(row=i, column=4).value = None
                # camera_Pac.cell(row=i, column=5).value = None
                # camera_Pac.cell(row=i, column=6).value = None
                print("One more shit was deleted ", camera_Pac.cell(row=i, column=5).value,
                      camera_Pac.cell(row=i, column=6).value)
                data_base.save("dataBase.xlsx")
            else:
                j += 1
        i += 1


# Функции для подсчета Intersection over Union (IoU)
def calculate_iou(box, camera):
    # Считаем IoU
    t = 0
    totalIOU = 0
    maxIOU = 0
    flag = 1
    async with _db_pool.acquire() as conn:
        lenth = await conn.fetchval(f"SELECT COUNT(*) FROM {camera}")
    x = 1
    for i in range(lenth):
        place = Place(i)
        place.load(camera)
        y1 = np.maximum(box[0], place.x)
        y2 = np.minimum(box[2] + box[0], place.h + place.x)
        x1 = np.maximum(box[1], place.y)
        x2 = np.minimum(box[3] + box[1], place.w + place.y)
        intersection = np.maximum(x2 - x1, 0) * np.maximum(y2 - y1, 0)
        union = box[2] * box[3] + place.h * place.w - intersection
        iou = intersection / union
        totalIOU += iou
        if iou > 0.6:
            place.finde_midle(box)
            tO = time.time()
            place.sub = iou
            tN = time.time()
            t += tO - tN
            flag = 0
        if iou > 0.15:
            place.free = False
            flag = 0
        else:
            if place.sub + iou > 0.15:
                place.free = False
            place.sub += iou
        place.load(camera)

    if flag:
        place = Place(lenth + x)
        x += 1
        place.set(box, 0, 0, False, 0)
        place.push(camera)
    return iou


# Функция для расчета персечения всех со всеми через IoU
def compute_overlaps(boxes, camera):
    setIOU(camera)
    init_db()
    now_all_space_free(camera)
    for box in boxes:
        place = Place(-1)
        place.set(box)
        calculate_iou(box, camera)
    return


def setIOU(camera):
    async with _db_pool.acquire() as conn:
        await conn.execute(
            f"UPDATE {camera} SET sub = 0"
        )


def draw_data(image_to_process, boxes, parking_color=(0, 255, 0)):
    color = parking_color
    width = 2
    for i in boxes:
        for place in i:
            x, y, w, h = int(place.x), int(place.y), int(place.w), int(place.h)
            start = (x, y)
            end = (x + w, y + h)
            if place.free and place.confidence > 5:
                color = (0, 255, 0)

            elif place.free and place.confidence <= 5:
                color = (0, 165, 255)
            else:
                color = (255, 0, 0)
            image_to_process = cv2.rectangle(image_to_process, start, end, color, width)
    return image_to_process"""
