import asyncio

from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor
import functools
import app
from ultralytics import YOLO
import time

camerasFile = load_workbook("Cameras.xlsx")
model = YOLO('YOLO-weights/best_v36.pt') # Здесь менать модель для нейронки

def getAddressesList():
    return camerasFile.get_sheet_names()



def getAddressesString():
    addresses = camerasFile.get_sheet_names()
    result = ''
    for address in addresses:
        result += f", {address}"
    return result


def getAddresses():
    inData = camerasFile.get_sheet_names()
    outData = []
    for i in range(len(inData)):
        outData.append([inData[i]])
    return outData


# async def detAnalysisAddresses(address):
    """
    Последовательно
    """
#     freeSpace = []
#     freeSpaceImg = []
#     cameras = {}
#     camerasList = []
#     answer = []
#     for i in range(1, camerasFile[address].max_row + 1):
#         camerasList.append(camerasFile[address].cell(row=i, column=1).value)
#         cameras[camerasFile[address].cell(row=i, column=1).value] = camerasFile[address].cell(row=i, column=2).value
#
#     for camera in camerasList:
#         detectResult = await app.detect(camera, cameras[camera], model)
#         answer.append(detectResult)
#         freeSpaceImg.append(detectResult[0])
#         freeSpace.append(detectResult[1])
#
#     return answer


async def detAnalysisAddresses(address):
    """
    Асинхронная функция для параллельного анализа видео с нескольких камер,
    связанных с указанным адресом.

    Args:
        address (str): Адрес/ключ для поиска камер в camerasFile

    Returns:
        tuple: Кортеж с результатами анализа:
            - answer: полные результаты детекции
            - freeSpaceImg: изображения со свободными местами
            - freeSpace: данные о свободных местах
    """

    # ===== БЛОК 1: Подготовка данных о камерах =====
    # Создаем словарь и список для хранения информации о камерах
    cameras = {}  # Словарь {название_камеры: путь_к_видео}
    camerasList = []  # Список камер

    # Перебираем все строки в файле камер для данного адреса
    for i in range(1, camerasFile[address].max_row + 1):
        # Получаем название камеры из первого столбца
        camera = camerasFile[address].cell(row=i, column=1).value
        # Добавляем камеру в список
        camerasList.append(camera)
        # Сохраняем путь к видео из второго столбца в словарь
        cameras[camera] = camerasFile[address].cell(row=i, column=2).value

    # ===== БЛОК 2: Настройка параллельного выполнения =====
    # Создаем частичную функцию с фиксированными аргументами
    detect_partial = functools.partial(run_detect_in_process, cameras=cameras)

    # ===== БЛОК 3: Параллельное выполнение обработки камер =====
    x = time.time()

    # Создаем пул процессов для параллельного выполнения
    with ProcessPoolExecutor() as executor:
        # Запускаем задачи для каждой камеры в пуле процессов
        futures = [
            executor.submit(detect_partial, camera)
            for camera in camerasList
        ]

        # ===== БЛОК 4: Сбор и обработка результатов =====
        # Инициализируем списки для результатов
        answer = []  # Для полных результатов
        freeSpaceImg = []  # Для изображений
        freeSpace = []  # Для данных о свободных местах

        # Обрабатываем результаты по мере их готовности
        for future in futures:
            # Получаем результат выполнения задачи
            detectResult = future.result()

            # Сохраняем результаты в соответствующие списки
            answer.append(detectResult)  # Полный результат
            freeSpaceImg.append(detectResult[0])  # Первый элемент - изображение
            freeSpace.append(detectResult[1])  # Второй элемент - данные о местах

    # Возвращаем собранные результаты
    return answer

def run_detect_in_process(camera, cameras):
    """Функция-обертка для запуска в отдельном процессе"""
    print(camera, cameras)
    return asyncio.run(app.detect(camera, cameras[camera], model))


getAddressesString()
